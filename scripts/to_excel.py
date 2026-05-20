#!/usr/bin/env python3
"""
测试用例导出 Excel 工具（v0.2）

基于 docs/specs/测试用例导入规范.md 和 docs/specs/测试用例文本协议.md 规范

输出格式严格遵循导入规范定义的字段：
- 必填：优先级、用例标题、操作步骤、预期结果、测试类型
- 可选：一级分组、二级分组、前置条件、是否反向用例、备注

用法：
    python to_excel.py test-case/all_cases.md -o output.xlsx
    python to_excel.py test-case/ -o output.xlsx
"""

import sys
import argparse
from pathlib import Path
from typing import List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 需要安装 openpyxl: `uv add openpyxl`")
    sys.exit(1)

from parse_text_protocol import parse_test_cases, TestCase, XMindFormatParser


# Excel 列配置（严格遵循 测试用例导入规范.md）
EXCEL_COLUMNS = [
    # 必填字段
    {"field": "优先级", "width": 8, "center": True},
    {"field": "用例标题", "width": 40},
    {"field": "操作步骤", "width": 50, "wrap": True},
    {"field": "预期结果", "width": 40, "wrap": True},
    {"field": "测试类型", "width": 12, "center": True},
    # 可选字段
    {"field": "前置条件", "width": 25},
    {"field": "是否反向用例", "width": 12, "center": True},
    {"field": "一级分组", "width": 18},
    {"field": "二级分组", "width": 18},
    {"field": "备注", "width": 20},
]


class ExcelExporter:
    """Excel 导出器"""

    def __init__(self, sheet_name: str = "测试用例"):
        self.sheet_name = sheet_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

    def export(self, rows: List[Dict], output_file: Path):
        """导出测试用例到 Excel"""
        self._write_header()
        self._write_data(rows)
        self._apply_styles()
        self.wb.save(output_file)
        print(f"✅ 导出成功: {output_file}")
        print(f"   总计 {len(rows)} 个用例")

    def _write_header(self):
        """写入表头"""
        for col_idx, col_config in enumerate(EXCEL_COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = col_config["field"]
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            self.ws.column_dimensions[get_column_letter(col_idx)].width = col_config["width"]

    def _write_data(self, rows: List[Dict]):
        """写入数据"""
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_config in enumerate(EXCEL_COLUMNS, 1):
                field = col_config["field"]
                value = row_data.get(field, "")
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                if col_config.get("center"):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_config.get("wrap"):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")

    def _apply_styles(self):
        """应用边框样式"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, 
                                      min_col=1, max_col=len(EXCEL_COLUMNS)):
            for cell in row:
                cell.border = thin_border


def case_to_row(case: TestCase, item: str = "", point: str = "") -> Dict:
    """将测试用例对象转换为 Excel 行数据"""
    steps_str = "\n".join(f"{i}. {step}" for i, step in enumerate(case.steps, 1))
    expected_str = "\n".join(
        f"{i}. {exp}" for i, exp in enumerate(case.expected, 1) if exp
    )

    return {
        "优先级": case.priority or "",
        "用例标题": case.title,
        "操作步骤": steps_str,
        "预期结果": expected_str,
        "测试类型": case.test_type or "功能测试",
        "前置条件": case.precondition,
        "是否反向用例": "否",
        "一级分组": item,
        "二级分组": point,
        "备注": case.note,
    }


def collect_from_file(file_path: Path, item: str = "", point: str = "") -> List[Dict]:
    """从单个文件收集用例"""
    rows = []
    try:
        cases = parse_test_cases(file_path, strict=False)
        
        # 如果未指定 item/point，从路径推断
        if not item and not point:
            parts = file_path.parts
            if len(parts) >= 2:
                item = parts[-2]
                point = file_path.stem
            else:
                point = file_path.stem

        for case in cases:
            rows.append(case_to_row(case, item, point))
        
        if cases:
            print(f"✅ {item}/{point}: {len(cases)} 个用例")

    except Exception as e:
        print(f"⚠️  解析失败 {file_path}: {e}")

    return rows


def collect_from_directory(dir_path: Path) -> List[Dict]:
    """从目录收集所有用例"""
    all_rows = []

    for item_dir in dir_path.iterdir():
        if not item_dir.is_dir() or item_dir.name.startswith('.'):
            continue

        item_name = item_dir.name
        for md_file in item_dir.glob('*.md'):
            if md_file.name in ['plan.md', 'all_cases.md']:
                continue
            point_name = md_file.stem
            rows = collect_from_file(md_file, item_name, point_name)
            all_rows.extend(rows)

    return all_rows


def collect_from_merged_file(file_path: Path) -> List[Dict]:
    """
    从合并文件（all_cases.md）收集用例

    合并文件格式（XMind 导入格式）：
    # 根节点（系统名，不被使用）
    ## 模块名（ITEM，一级分组）
    ### 测试点名（POINT，二级分组）
    #### tc-P1-功能测试: 用例标题
    * pc: 前置条件
    * 步骤1
        * 预期结果1
    """
    import re
    from parse_text_protocol import XMindFormatParser

    rows = []
    current_item = ""
    current_point = ""

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        i = 0
        while i < len(lines):
            line = lines[i].rstrip()
            stripped = line.strip()

            # ## 模块名（ITEM）— 不匹配 ### 或 ####
            if re.match(r'^##\s', stripped) and not re.match(r'^###', stripped):
                current_item = stripped[2:].strip()
                i += 1
                continue

            # ### 测试点名（POINT）— 不匹配 ####
            if re.match(r'^###\s', stripped) and not re.match(r'^####', stripped):
                current_point = stripped[3:].strip()
                i += 1
                continue

            # #### tc... 用例节点
            if re.match(r'^####\s+tc', stripped):
                case_lines = [line]
                i += 1
                while i < len(lines):
                    next_stripped = lines[i].strip()
                    if re.match(r'^#{1,4}\s', next_stripped):
                        break
                    case_lines.append(lines[i].rstrip())
                    i += 1

                parser = XMindFormatParser(strict=False)
                cases = parser.parse_content('\n'.join(case_lines))
                for case in cases:
                    rows.append(case_to_row(case, current_item, current_point))
                continue

            i += 1

        print(f"✅ 从 {file_path.name} 解析: {len(rows)} 个用例")

    except Exception as e:
        print(f"⚠️  解析失败 {file_path}: {e}")

    return rows


def main():
    parser = argparse.ArgumentParser(description="将测试用例导出为 Excel 格式")
    parser.add_argument("input", help="输入文件或目录路径")
    parser.add_argument("-o", "--output", required=True, help="输出 Excel 文件路径")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"❌ 路径不存在: {input_path}")
        sys.exit(1)

    print(f"开始导出测试用例: {input_path}\n")

    # 收集用例
    if input_path.is_file():
        if input_path.name == 'all_cases.md':
            rows = collect_from_merged_file(input_path)
        else:
            rows = collect_from_file(input_path)
    elif input_path.is_dir():
        rows = collect_from_directory(input_path)
    else:
        print(f"❌ 无效的路径: {input_path}")
        sys.exit(1)

    if not rows:
        print("⚠️  没有找到任何测试用例")
        sys.exit(0)

    # 导出
    exporter = ExcelExporter()
    exporter.export(rows, output_path)


if __name__ == "__main__":
    main()
